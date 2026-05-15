from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.sced_fields import SUPPLEMENT_3_FIELDS

PDF_DIR = ROOT / "input"
XLSX_PATH = ROOT / "data" / "Supplement 3. Case and study characteristics.xlsx"
OUT_PATH = ROOT / "data" / "sced_gold.jsonl"


def normalize(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def clean_value(value: object) -> object | None:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.upper() == "NA":
            return None
        return stripped
    return value


def build_pdf_lookup() -> dict[str, str]:
    pdfs = sorted(
        p.name for p in PDF_DIR.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"
    )
    lookup: dict[str, str] = {}
    for pdf in pdfs:
        stem = Path(pdf).stem
        norm = normalize(stem)
        lookup[norm] = pdf
    return lookup


def match_pdf(study: str, pdf_lookup: dict[str, str]) -> str | None:
    study_norm = normalize(study)
    exact = pdf_lookup.get(study_norm)
    if exact:
        return exact

    candidates = [pdf for norm, pdf in pdf_lookup.items() if study_norm in norm or norm in study_norm]
    if len(candidates) == 1:
        return candidates[0]
    if candidates:
        return sorted(candidates, key=len)[0]

    year_match = re.search(r"(19|20)\d{2}", study)
    if year_match:
        year = year_match.group(0)
        author_part = study[: year_match.start()]
        author_tokens = [
            token.lower()
            for token in re.findall(r"[A-Za-z]+", author_part)
            if len(token) > 2
        ]
        author_year_matches: list[tuple[int, str]] = []
        for pdf_name in pdf_lookup.values():
            pdf_lower = pdf_name.lower()
            if year not in pdf_lower:
                continue
            matched_tokens = sum(
                1
                for token in author_tokens
                if re.search(rf"\b{re.escape(token)}\b", pdf_lower)
            )
            if matched_tokens:
                author_year_matches.append((matched_tokens, pdf_name))

        if author_year_matches:
            best_score = max(score for score, _ in author_year_matches)
            best_matches = [
                pdf_name
                for score, pdf_name in author_year_matches
                if score == best_score
            ]
            if len(best_matches) == 1:
                return best_matches[0]

    return None


def collapse(values: list[object]) -> object | None:
    cleaned: list[object] = []
    seen: set[str] = set()
    for value in values:
        value = clean_value(value)
        if value is None:
            continue
        marker = repr(value)
        if marker in seen:
            continue
        seen.add(marker)
        cleaned.append(value)
    if not cleaned:
        return None
    if len(cleaned) == 1:
        return cleaned[0]
    return cleaned


def load_excel_records() -> dict[str, dict[str, object]]:
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    current_study: str | None = None
    rows_by_study: dict[str, list[dict[str, object]]] = {}
    empty_row_run = 0

    for row_idx in range(2, ws.max_row + 1):
        row_values = [
            ws.cell(row_idx, field.excel_column).value
            for field in SUPPLEMENT_3_FIELDS
        ]
        if all(clean_value(value) is None for value in row_values):
            empty_row_run += 1
            if empty_row_run >= 50:
                break
            continue

        empty_row_run = 0
        study = clean_value(row_values[0])
        if study is not None:
            current_study = str(study)
            rows_by_study.setdefault(current_study, [])
        if current_study is None:
            continue

        rows_by_study[current_study].append(
            {
                field.name: value
                for field, value in zip(SUPPLEMENT_3_FIELDS, row_values)
            }
        )

    records: dict[str, dict[str, object]] = {}
    for study, rows in rows_by_study.items():
        records[study] = {
            field.name: collapse([row[field.name] for row in rows])
            for field in SUPPLEMENT_3_FIELDS
        }
    return records


def main() -> None:
    pdf_lookup = build_pdf_lookup()
    study_records = load_excel_records()
    per_pdf: dict[str, dict[str, object]] = {}
    unmatched_studies: set[str] = set()
    for study, record in study_records.items():
        pdf_name = match_pdf(study, pdf_lookup)
        if not pdf_name:
            unmatched_studies.add(study)
            continue
        per_pdf[pdf_name] = record

    with OUT_PATH.open("w", encoding="utf-8") as handle:
        for pdf_name in sorted(per_pdf):
            handle.write(json.dumps({"pdf": pdf_name, **per_pdf[pdf_name]}, ensure_ascii=False) + "\n")

    print(f"Wrote {len(per_pdf)} records to {OUT_PATH}")
    if unmatched_studies:
        print("Unmatched studies:")
        for study in sorted(unmatched_studies):
            print(f"- {study}")


if __name__ == "__main__":
    main()
